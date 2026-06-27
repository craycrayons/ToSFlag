"""Export a layman-readable per-clause report from the ToSFlag pipeline.

This is a presentation layer over the EXACT Lap-1 pipeline in run.py. It does not
reimplement anything: it imports tosflag.data and tosflag.model, fits on train,
chooses the threshold on validation via the cost asymmetry, then writes one row
per TEST clause that a non-technical reader can open in Excel/Sheets and act on.

    clause_text | true_label | severity | flagged | risk_score |
    outcome | is_nonclause | why_flagged

Outputs both .csv (opens anywhere) and .xlsx (shaded by severity, filterable).

Run (mirrors run.py exactly):
    python scripts/export_report.py
    python scripts/export_report.py --local data/train.parquet data/validation.parquet data/test.parquet
    python scripts/export_report.py --split all      # report on train+val+test, not just test
    python scripts/export_report.py --csv-only

CPU-only. No GPU, no torch, no legal-bert involved -- this is the TF-IDF path.

Honest scope (stated so the report cannot be mistaken for more than it is):
  - 'severity' is the dataset's own 3-class label (clearly_fair / potentially_
    unfair / clearly_unfair), NOT a model output. It tells a reader how the
    ground truth rates a clause. The model produces 'risk_score' + 'flagged'.
  - 'why_flagged' lists the highest-weight unfair-direction tokens the model
    saw in the clause. It is a transparency aid, not a legal classification.
  - 'is_nonclause' marks header/stub label-noise (per data.is_nonclause), so a
    reader can discount those rows -- the same correction run.py applies to the
    recall number.
  - The flagged/score columns come from the model's decision on the held-out
    TEST split (no peeking): this is a real generalisation view, not in-sample.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from tosflag import data, model  # noqa: E402

REPORTS = Path(__file__).resolve().parents[1] / "reports"


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--local", nargs=3, metavar=("TRAIN", "VAL", "TEST"))
    ap.add_argument("--split", choices=["test", "all"], default="test",
                    help="which split(s) to write rows for (default: test, the held-out view)")
    ap.add_argument("--out", default="clause_report", help="output basename")
    ap.add_argument("--csv-only", action="store_true")
    args = ap.parse_args()

    frames = (
        data.load_local_frames(
            {"train": args.local[0], "validation": args.local[1], "test": args.local[2]}
        )
        if args.local
        else data.load_frames()
    )
    REPORTS.mkdir(exist_ok=True)
    train, val, test = frames["train"], frames["validation"], frames["test"]

    # Fit on train, choose the operating point on validation -- identical to run.py.
    pipe = model.build_pipeline().fit(train["text"], train["y_binary"])
    val_scores = pipe.predict_proba(val["text"])[:, 1]
    op = model.choose_threshold(val["y_binary"].values, val_scores)
    print(f"Cost ratio FN:FP = {model.FN_TO_FP_COST_RATIO}:1  ->  threshold = {op.threshold:.3f}")

    # Precompute feature names + coefficients once for the why_flagged column.
    names = _build_feature_names(pipe)
    try:
        coefs = pipe.named_steps["clf"].coef_[0]
    except Exception:
        coefs, names = None, None

    # Which frames to write rows for.
    targets = {"test": test}
    if args.split == "all":
        targets = {"train": train, "validation": val, "test": test}

    all_rows = []
    for split_name, df in targets.items():
        scores = pipe.predict_proba(df["text"])[:, 1]
        all_rows.extend(_rows_for_split(split_name, df, scores, op.threshold, pipe, names, coefs))

    report = pd.DataFrame(all_rows).sort_values(
        ["split", "risk_score"], ascending=[True, False]
    )

    # Print the same honest recall summary run.py reports, so this script is
    # self-documenting about what the numbers mean.
    _print_summary(test, pipe, op.threshold)

    csv_path = REPORTS / f"{args.out}.csv"
    report.to_csv(csv_path, index=False, encoding="utf-8")
    print(f"\nWrote {csv_path}")

    if not args.csv_only:
        xlsx_path = REPORTS / f"{args.out}.xlsx"
        _write_xlsx(report, xlsx_path, op, test, pipe)
        print(f"Wrote {xlsx_path}")


def _rows_for_split(split_name, df, scores, threshold, pipe, names, coefs):
    rows = []
    y = df["y_binary"].values
    levels = df["unfairness_level"].values
    nonclause = df["is_nonclause"].values
    texts = df["text"].values
    flagged = (scores >= threshold).astype(int)
    for txt, true_l, lvl, nc, sc, fl in zip(texts, y, levels, nonclause, scores, flagged):
        if fl and true_l == 1:
            outcome = "caught"
        elif fl and true_l == 0:
            outcome = "false_flag"
        elif not fl and true_l == 1:
            outcome = "MISSED"
        else:
            outcome = "correct_pass"
        rows.append({
            "split": split_name,
            "clause_text": txt,
            "true_label": "unfair" if true_l == 1 else "fair",
            "severity": lvl,                       # dataset's own 3-class rating
            "flagged": "YES" if fl else "no",
            "risk_score": round(float(sc), 3),
            "outcome": outcome,
            "is_nonclause": "yes" if nc else "",
            "why_flagged": _why(txt, pipe, names, coefs) if fl else "",
        })
    return rows


def _build_feature_names(pipe):
    """Assemble the FeatureUnion's column names in concatenation order.

    FeatureUnion.get_feature_names_out() raises because StructuralFeatures has no
    such method, so we pull names per sub-transformer and pad the 4 structural
    columns with placeholders. Returns None if the structure isn't the expected
    TF-IDF union (e.g. a future scorer swap), so callers degrade gracefully.
    """
    try:
        feats = pipe.named_steps["features"]
        out = []
        for name, trans in feats.transformer_list:
            if hasattr(trans, "get_feature_names_out"):
                out.extend(list(trans.get_feature_names_out()))
            elif hasattr(trans, "vec"):  # _CapsTfidf wraps .vec
                out.extend(list(trans.vec.get_feature_names_out()))
            else:                        # StructuralFeatures: 4 non-token columns
                out.extend([f"__struct_{i}" for i in range(4)])
        return out
    except Exception:
        return None


def _why(text, pipe, names, coefs, k=4):
    """Top-k high-weight unfair-direction tokens present in the clause. Plain
    transparency on why a row lit up; '' if names unavailable."""
    if not names or coefs is None:
        return ""
    try:
        row = pipe.named_steps["features"].transform([str(text)])
        present = row.nonzero()[1]
        scored = sorted(
            ((coefs[i], names[i]) for i in present
             if i < len(coefs) and coefs[i] > 0 and not names[i].startswith("__struct_")),
            reverse=True,
        )
        # de-dup while preserving order (the two TF-IDF blocks can repeat a token)
        seen, terms = set(), []
        for _, n in scored:
            if n not in seen:
                seen.add(n); terms.append(n)
            if len(terms) >= k:
                break
        return ", ".join(terms)
    except Exception:
        return ""


def _print_summary(test, pipe, threshold):
    """Reproduce run.py's all-items vs real-clauses recall, so a reader of this
    report sees the honest (stub-excluded) number, not just the headline."""
    scores = pipe.predict_proba(test["text"])[:, 1]
    y = test["y_binary"].values
    all_m = model.metrics_at(y, scores, threshold)
    mask = ~test["is_nonclause"].values
    clause_m = model.metrics_at(y[mask], scores[mask], threshold)
    n_stub_unfair = int((test["is_nonclause"].values & (y == 1)).sum())
    print(f"\nTEST recall (all items):    {all_m.recall:.3f}  "
          f"precision {all_m.precision:.3f}  flagged {all_m.flagged_rate:.1%}")
    print(f"TEST recall (real clauses): {clause_m.recall:.3f}  "
          f"(honest number; {n_stub_unfair} 'unfair' stubs excluded as label noise)")


def _write_xlsx(report, path, op, test, pipe):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Summary sheet ---
    s = wb.active
    s.title = "Summary"
    s["A1"] = "ToSFlag - clause report"
    s["A1"].font = Font(bold=True, size=14)

    scores = pipe.predict_proba(test["text"])[:, 1]
    y = test["y_binary"].values
    all_m = model.metrics_at(y, scores, op.threshold)
    mask = ~test["is_nonclause"].values
    clause_m = model.metrics_at(y[mask], scores[mask], op.threshold)
    n_stub_unfair = int((test["is_nonclause"].values & (y == 1)).sum())

    summary = [
        ("Cost ratio FN:FP", f"{model.FN_TO_FP_COST_RATIO}:1"),
        ("Operating threshold", round(op.threshold, 3)),
        ("TEST recall (all items)", round(all_m.recall, 3)),
        ("TEST recall (real clauses, honest)", round(clause_m.recall, 3)),
        ("'Unfair' stubs excluded as label noise", n_stub_unfair),
        ("TEST precision", round(all_m.precision, 3)),
        ("TEST flagged rate", f"{all_m.flagged_rate:.1%}"),
    ]
    bold = Font(bold=True)
    for i, (k, v) in enumerate(summary, start=3):
        s[f"A{i}"] = k; s[f"B{i}"] = v; s[f"A{i}"].font = bold
    s["A12"] = ("'severity' is the dataset's own 3-class label, not a model output. "
                "'risk_score'/'flagged' are the model's decision on held-out test data.")
    s["A13"] = ("Rows marked is_nonclause=yes are header/stub label noise; discount them. "
                "outcome=MISSED are the costly false negatives.")
    s.column_dimensions["A"].width = 38
    s.column_dimensions["B"].width = 16

    # --- Clauses sheet ---
    c = wb.create_sheet("Clauses")
    cols = list(report.columns)
    c.append(cols)
    for j in range(1, len(cols) + 1):
        cell = c.cell(row=1, column=j)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="top")

    # Shade by the dataset severity so a reader sees ground-truth risk at a glance.
    sev_fill = {
        "clearly_unfair": PatternFill("solid", start_color="F4CCCC"),     # red
        "potentially_unfair": PatternFill("solid", start_color="FFF2CC"),  # amber
        "clearly_fair": PatternFill("solid", start_color="FFFFFF"),       # white
    }
    miss_font = Font(bold=True, color="CC0000")
    out_idx = cols.index("outcome")
    sev_idx = cols.index("severity")
    for _, r in report.iterrows():
        c.append([r[col] for col in cols])
        fill = sev_fill.get(r["severity"])
        if fill:
            for j in range(1, len(cols) + 1):
                cl = c.cell(row=c.max_row, column=j)
                cl.fill = fill
                cl.alignment = Alignment(vertical="top", wrap_text=(cols[j-1] == "clause_text"))
        if r["outcome"] == "MISSED":  # make the costly errors jump out
            c.cell(row=c.max_row, column=out_idx + 1).font = miss_font

    widths = {"split": 11, "clause_text": 78, "true_label": 11, "severity": 18,
              "flagged": 9, "risk_score": 11, "outcome": 13, "is_nonclause": 12,
              "why_flagged": 34}
    for j, name in enumerate(cols, start=1):
        c.column_dimensions[get_column_letter(j)].width = widths.get(name, 14)
    c.freeze_panes = "A2"
    c.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{c.max_row}"

    wb.save(path)


if __name__ == "__main__":
    main()
