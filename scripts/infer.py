"""Score bring-your-own ToS clauses with the frozen ToSFlag model.

The inference path: takes clauses you already split (one per line), scores each
with the committed TF-IDF model + threshold from train_and_save.py, and writes a
report with NO ground-truth columns -- unseen text has no answer key, so there is
no true_label / outcome / severity, only what the model produces:

    clause | flagged | risk_score | why_flagged

Input is one clause per line. Blank lines are ignored. Segmentation
(wall-of-text -> clauses) is deliberately OUT of scope; you split first.

Run:
    python scripts/infer.py my_clauses.txt
    cat my_clauses.txt | python scripts/infer.py          # piped input
    python scripts/infer.py my_clauses.txt --out my_report --csv-only

Outputs (to reports/, basename --out, default 'byo_report'):
    <out>.csv    -- complete, opens anywhere (the machine/archive copy)
    <out>.xlsx   -- colored for reading: rows shaded by risk band, worst first

CPU-only. Loads models/tosflag_tfidf.joblib; no HuggingFace, no torch, no
training at inference time.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import joblib
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))  # so we can import export_report

from tosflag import model  # noqa: E402  (kept importable for unpickling the custom classes)
from export_report import _build_feature_names, _why  # noqa: E402  -- reuse exact why_flagged logic

MODELS = ROOT / "models"
REPORTS = ROOT / "reports"
MODEL_PATH = MODELS / "tosflag_tfidf.joblib"
META_PATH = MODELS / "tosflag_meta.json"


def read_clauses(path: str | None) -> list[str]:
    """One clause per line. From `path` if given, else stdin. Blanks dropped."""
    if path:
        raw = Path(path).read_text(encoding="utf-8")
    else:
        if sys.stdin.isatty():
            sys.exit("No input. Pass a file (python infer.py clauses.txt) "
                     "or pipe text (cat clauses.txt | python infer.py).")
        raw = sys.stdin.read()
    return [ln.strip() for ln in raw.splitlines() if ln.strip()]


def load_model_and_threshold():
    if not MODEL_PATH.exists() or not META_PATH.exists():
        sys.exit(f"Missing frozen model. Run scripts/train_and_save.py first "
                 f"(expected {MODEL_PATH} and {META_PATH}).")
    pipe = joblib.load(MODEL_PATH)
    meta = json.loads(META_PATH.read_text(encoding="utf-8"))
    return pipe, float(meta["threshold"]), meta


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("input", nargs="?", help="text file, one clause per line (omit to read stdin)")
    ap.add_argument("--out", default="byo_report", help="output basename (default: byo_report)")
    ap.add_argument("--csv-only", action="store_true")
    args = ap.parse_args()

    clauses = read_clauses(args.input)
    if not clauses:
        sys.exit("No non-empty lines to score.")

    pipe, threshold, meta = load_model_and_threshold()
    scores = pipe.predict_proba(clauses)[:, 1]

    # Same why_flagged transparency aid as export_report, by import not copy.
    names = _build_feature_names(pipe)
    try:
        coefs = pipe.named_steps["clf"].coef_[0]
    except Exception:
        coefs, names = None, None

    rows = []
    for txt, sc in zip(clauses, scores):
        flagged = sc >= threshold
        rows.append({
            "clause": txt,
            "flagged": "YES" if flagged else "no",
            "risk_score": round(float(sc), 3),
            "why_flagged": _why(txt, pipe, names, coefs) if flagged else "",
        })
    report = pd.DataFrame(rows).sort_values("risk_score", ascending=False).reset_index(drop=True)

    REPORTS.mkdir(exist_ok=True)
    n_flag = int((report["flagged"] == "YES").sum())
    print(f"Scored {len(report)} clauses at threshold {threshold:.3f} "
          f"(FN:FP={meta.get('fn_fp_cost')}:1, model trained {meta.get('trained_at')}). "
          f"Flagged {n_flag}/{len(report)}.")

    csv_path = REPORTS / f"{args.out}.csv"
    report.to_csv(csv_path, index=False, encoding="utf-8")
    print(f"Wrote {csv_path}")

    if not args.csv_only:
        xlsx_path = REPORTS / f"{args.out}.xlsx"
        _write_xlsx(report, xlsx_path, threshold, meta)
        print(f"Wrote {xlsx_path}")


def _write_xlsx(report: pd.DataFrame, path: Path, threshold: float, meta: dict) -> None:
    """Colored for human reading. No severity ground truth on unseen text, so
    rows are shaded by the model's own risk_score band instead -- the only
    signal we have. Eye still lands on the worst clauses first."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Summary sheet ---
    s = wb.active
    s.title = "Summary"
    s["A1"] = "ToSFlag - bring-your-own-clauses report"
    s["A1"].font = Font(bold=True, size=14)
    n_flag = int((report["flagged"] == "YES").sum())
    summary = [
        ("Clauses scored", len(report)),
        ("Flagged for review", n_flag),
        ("Operating threshold", round(threshold, 3)),
        ("Cost ratio FN:FP", f"{meta.get('fn_fp_cost')}:1"),
        ("Model trained", meta.get("trained_at")),
        ("scikit-learn version", meta.get("sklearn_version")),
    ]
    bold = Font(bold=True)
    for i, (k, v) in enumerate(summary, start=3):
        s[f"A{i}"] = k; s[f"B{i}"] = v; s[f"A{i}"].font = bold
    s["A11"] = ("risk_score and flagged are the model's decision on UNSEEN text. "
                "There is no true_label/severity here -- pasted clauses have no answer key.")
    s["A12"] = ("Rows shaded by risk band: red = high score, amber = above threshold, "
                "white = below threshold (model thinks fair). why_flagged shows the "
                "highest-weight tokens behind a flag -- a transparency aid, not legal advice.")
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 22

    # --- Clauses sheet ---
    c = wb.create_sheet("Clauses")
    cols = list(report.columns)
    c.append(cols)
    for j in range(1, len(cols) + 1):
        cell = c.cell(row=1, column=j)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="top")

    # Risk bands off the committed threshold. Midpoint between threshold and 1.0
    # splits "flagged" into amber (just over the line) vs red (high confidence).
    red_cut = threshold + (1.0 - threshold) / 2.0
    red = PatternFill("solid", start_color="F4CCCC")
    amber = PatternFill("solid", start_color="FFF2CC")
    white = PatternFill("solid", start_color="FFFFFF")

    for _, r in report.iterrows():
        c.append([r[col] for col in cols])
        sc = r["risk_score"]
        fill = red if sc >= red_cut else amber if sc >= threshold else white
        for j in range(1, len(cols) + 1):
            cl = c.cell(row=c.max_row, column=j)
            cl.fill = fill
            cl.alignment = Alignment(vertical="top", wrap_text=(cols[j-1] == "clause"))

    widths = {"clause": 90, "flagged": 9, "risk_score": 11, "why_flagged": 36}
    for j, name in enumerate(cols, start=1):
        c.column_dimensions[get_column_letter(j)].width = widths.get(name, 14)
    c.freeze_panes = "A2"
    c.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{c.max_row}"

    wb.save(path)


if __name__ == "__main__":
    main()
